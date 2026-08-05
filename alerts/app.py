"""
alerts/app.py — Dashboard de Alertas de Transcripción TV.
Flask app: auth, búsquedas, coincidencias, reportes, panel admin.
"""
import json
import os
import re
import sqlite3
import unicodedata
from datetime    import date, datetime, timedelta
from functools   import wraps, lru_cache
from pathlib     import Path

from flask              import (Flask, g, flash, jsonify, redirect,
                                render_template, request, session, url_for,
                                send_file)
from markupsafe         import Markup, escape as html_escape
from werkzeug.security  import check_password_hash, generate_password_hash

BASE_DIR  = Path(__file__).parent.parent
ALERTS_DB = BASE_DIR / 'alerts.db'
TRANS_DB  = BASE_DIR / 'transcriptions.db'


# ── Resaltado de keywords ─────────────────────────────────────────────────────
@lru_cache(maxsize=8192)
def _strip_acc(s: str) -> str:
    s = unicodedata.normalize('NFD', s.lower())
    return ''.join(c for c in s if unicodedata.category(c) != 'Mn')

@lru_cache(maxsize=8192)
def _phonetic(s: str) -> str:
    t = _strip_acc(s)
    t = re.sub(r'\bh', '', t)
    t = t.replace('v', 'b').replace('ll', 'y').replace('z', 's').replace('ck', 'k')
    t = re.sub(r'qu([ei])', r'k\1', t)
    t = re.sub(r'c([ei])', r's\1', t)
    t = re.sub(r'g([ei])', r'j\1', t)
    t = re.sub(r'x', 'ks', t)
    return t

CHUNK_SECONDS = 30   # debe coincidir con worker.py / transcriber.py

def _locate_keyword(text, keyword, phonetic=False, whole_word=False):
    """Encuentra el índice (basado en palabras) de la primera ocurrencia
    de la keyword (o frase de varias palabras). Devuelve (idx_word, total_words)
    o (None, total_words). Usa la misma lógica de match que _highlight
    (acento-insensitive y opcionalmente fonético).

    En modo whole_word, un token solo cuenta si la keyword aparece delimitada
    por separadores de palabra dentro de él (ej. "día" no debe casar con
    "diálogo") — se comprueba con límites \\w sobre el token normalizado, no
    con una lista fija de signos de puntuación, para cubrir cualquier
    puntuación pegada (paréntesis, comillas, etc.)."""
    if not text or not keyword:
        return None, 0
    words = text.split()
    n = len(words)
    if n == 0:
        return None, 0
    norm     = _phonetic if phonetic else _strip_acc
    kw_words = [norm(w) for w in keyword.split()]
    k = len(kw_words)
    if k == 0:
        return None, n

    if k == 1:
        kw_n    = kw_words[0]
        pattern = re.compile(r'(?<!\w)' + re.escape(kw_n) + r'(?!\w)') if whole_word else None
        for i, w in enumerate(words):
            w_n = norm(w)
            if whole_word:
                if pattern.search(w_n):
                    return i, n
            elif kw_n in w_n:
                return i, n
        return None, n

    # Frase de varias palabras: buscar la secuencia consecutiva de tokens.
    for i in range(n - k + 1):
        window = [norm(words[i + j]) for j in range(k)]
        if whole_word:
            if window == kw_words:
                return i, n
        elif all(kw_words[j] in window[j] for j in range(k)):
            return i, n
    return None, n

def _center_text(text, idx_word, words_each_side=30):
    """Recorta el texto centrado en la palabra match: N palabras antes y N después.
    Si no hay match, devuelve las primeras 2N+1 palabras."""
    if not text:
        return ''
    words = text.split()
    n = len(words)
    if n == 0:
        return ''
    if idx_word is None:
        idx_word = 0
    start = max(0, idx_word - words_each_side)
    end   = min(n, idx_word + words_each_side + 1)
    snippet = ' '.join(words[start:end])
    if start > 0:
        snippet = '… ' + snippet
    if end < n:
        snippet = snippet + ' …'
    return snippet

def _precise_timestamp(start_ts, idx_word, total_words, chunk_sec=CHUNK_SECONDS):
    """Estima el timestamp absoluto del momento exacto en que se mencionó la
    palabra dentro del chunk. start_ts es el inicio del audio (ya garantizado
    en worker.py al momento de captura, no de inferencia). Si la palabra está
    en la posición k de N, asumimos que se mencionó en (k+0.5)/N del chunk.
    Precisión típica ±2-3s para chunks de 30s."""
    if not start_ts or idx_word is None or not total_words:
        return start_ts
    try:
        # Soporta ISO con o sin milisegundos
        dt = datetime.fromisoformat(start_ts)
        offset = ((idx_word + 0.5) / total_words) * chunk_sec
        return (dt + timedelta(seconds=offset)).isoformat(sep=' ', timespec='seconds')
    except Exception:
        return start_ts

def _enrich_match(m, phonetic=False, whole_word=False):
    """Convierte una Row a dict y agrega: precise_timestamp y centered_text.
    `m` puede ser sqlite3.Row o dict."""
    md = dict(m)
    text = md.get('matched_text') or ''
    kw   = md.get('keyword') or ''
    idx, total = _locate_keyword(text, kw, phonetic=phonetic, whole_word=whole_word)
    md['precise_timestamp'] = _precise_timestamp(md.get('timestamp'), idx, total)
    md['centered_text']     = _center_text(text, idx, words_each_side=30)
    return md

def _highlight(text, keyword, phonetic=False, whole_word=False):
    """Resalta todas las ocurrencias de keyword en text con <mark>.
    Usa comparación sin acentos/mayúsculas; en modo fonético detecta
    palabras fonéticamente equivalentes aunque se escriban diferente.
    En modo whole_word exige límites de palabra (\\w) alrededor de la
    keyword, para no resaltar "día" dentro de "diálogo"."""
    if not text or not keyword:
        return Markup(html_escape(text or ''))
    text, keyword = str(text), str(keyword)
    if phonetic:
        ph_kw   = _phonetic(keyword)
        pattern = re.compile(r'(?<!\w)' + re.escape(ph_kw) + r'(?!\w)') if whole_word else None
        parts = re.split(r'(\s+)', text)
        out = []
        for p in parts:
            if not p.strip():
                out.append(str(html_escape(p)))
                continue
            p_ph    = _phonetic(p)
            matched = bool(pattern.search(p_ph)) if whole_word else (ph_kw in p_ph)
            if matched:
                out.append(f'<mark>{html_escape(p)}</mark>')
            else:
                out.append(str(html_escape(p)))
        return Markup(''.join(out))
    # Búsqueda exacta sin acentos/mayúsculas: regex case-insensitive sobre texto escapeado
    esc_text = str(html_escape(text))
    esc_kw   = re.escape(str(html_escape(keyword)))
    pattern  = (r'(?<!\w)' if whole_word else '') + esc_kw + (r'(?!\w)' if whole_word else '')
    result   = re.sub(pattern, lambda m: f'<mark>{m.group()}</mark>',
                      esc_text, flags=re.IGNORECASE)
    return Markup(result)


# ── DB schema ─────────────────────────────────────────────────────────────────
_SCHEMA = """
CREATE TABLE IF NOT EXISTS users (
    id            INTEGER PRIMARY KEY AUTOINCREMENT,
    name          TEXT    NOT NULL,
    email         TEXT    UNIQUE NOT NULL,
    password_hash TEXT    NOT NULL,
    role          TEXT    DEFAULT 'user',
    active        INTEGER DEFAULT 1,
    created_at    TEXT    DEFAULT (datetime('now','localtime'))
);
CREATE TABLE IF NOT EXISTS searches (
    id                INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id           INTEGER NOT NULL,
    name              TEXT    NOT NULL,
    keywords          TEXT    NOT NULL,
    phonetic          INTEGER DEFAULT 0,
    whole_word        INTEGER DEFAULT 0,
    date_start        TEXT    NOT NULL,
    date_end          TEXT    NOT NULL,
    status            TEXT    DEFAULT 'active',
    delivery_mode     TEXT    DEFAULT 'final',
    report_email      TEXT,
    last_daily_report TEXT,
    initialized       INTEGER DEFAULT 0,
    created_at        TEXT    DEFAULT (datetime('now','localtime')),
    FOREIGN KEY (user_id) REFERENCES users(id)
);
CREATE TABLE IF NOT EXISTS matches (
    id           INTEGER PRIMARY KEY AUTOINCREMENT,
    search_id    INTEGER NOT NULL,
    keyword      TEXT    NOT NULL,
    channel_id   INTEGER,
    channel_name TEXT,
    timestamp    TEXT,
    matched_text TEXT,
    emailed      INTEGER DEFAULT 0,
    found_at     TEXT    DEFAULT (datetime('now','localtime')),
    FOREIGN KEY (search_id) REFERENCES searches(id)
);
CREATE TABLE IF NOT EXISTS settings (
    key   TEXT PRIMARY KEY,
    value TEXT
);
CREATE INDEX IF NOT EXISTS idx_m_search ON matches(search_id);
CREATE INDEX IF NOT EXISTS idx_m_found  ON matches(found_at);
CREATE INDEX IF NOT EXISTS idx_s_user   ON searches(user_id);
CREATE INDEX IF NOT EXISTS idx_s_status ON searches(status, date_start, date_end);
"""


def _init_db():
    conn = sqlite3.connect(str(ALERTS_DB))
    conn.executescript(_SCHEMA)
    # Clave secreta persistente
    if not conn.execute("SELECT 1 FROM settings WHERE key='secret_key'").fetchone():
        conn.execute("INSERT INTO settings (key,value) VALUES ('secret_key',?)",
                     (os.urandom(32).hex(),))
    # Schema migrations
    for col, dfn in [
        ('notify_telegram',  'INTEGER DEFAULT 0'),   # searches table
        ('init_rows_done',   'INTEGER DEFAULT 0'),
        ('init_rows_total',  'INTEGER DEFAULT 0'),
        ('whole_word',       'INTEGER DEFAULT 0'),
    ]:
        try:
            conn.execute(f"ALTER TABLE searches ADD COLUMN {col} {dfn}")
            conn.commit()
        except Exception:
            pass
    try:
        conn.execute("ALTER TABLE users ADD COLUMN tg_chat_id TEXT DEFAULT ''")
        conn.commit()
    except Exception:
        pass
    # Eliminar duplicados antes de crear el índice único
    conn.execute("""
        DELETE FROM matches WHERE id NOT IN (
            SELECT MIN(id) FROM matches
            GROUP BY search_id, keyword, channel_id, timestamp
        )
    """)
    conn.execute("""
        CREATE UNIQUE INDEX IF NOT EXISTS idx_m_unique
        ON matches(search_id, keyword, channel_id, timestamp)
    """)
    conn.commit()
    conn.commit()
    conn.close()


def _get_secret_key() -> str:
    conn = sqlite3.connect(str(ALERTS_DB))
    row  = conn.execute("SELECT value FROM settings WHERE key='secret_key'").fetchone()
    conn.close()
    return row[0] if row else os.urandom(32).hex()


# ── App factory ───────────────────────────────────────────────────────────────
def create_app() -> Flask:
    _init_db()
    # Inicializar esquema EPG
    from alerts.epg import ensure_schema as _epg_schema
    _epg_conn = sqlite3.connect(str(ALERTS_DB))
    _epg_schema(_epg_conn)
    _epg_conn.close()

    app = Flask(__name__, template_folder='templates')
    app.secret_key = _get_secret_key()
    app.jinja_env.filters['fromjson']   = json.loads
    app.jinja_env.filters['highlight']  = _highlight

    # ── DB helpers ─────────────────────────────────────────────────
    def db():
        if 'db' not in g:
            g.db = sqlite3.connect(str(ALERTS_DB), timeout=10)
            g.db.row_factory = sqlite3.Row
            g.db.execute("PRAGMA journal_mode=WAL")
        return g.db

    @app.teardown_appcontext
    def _close_db(_=None):
        c = g.pop('db', None)
        if c:
            c.close()

    @app.after_request
    def _no_cache_html(resp):
        """Evita caché agresivo en páginas HTML dinámicas. Los assets estáticos
        siguen cacheándose normalmente."""
        ct = resp.headers.get('Content-Type', '')
        if 'text/html' in ct:
            resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
            resp.headers['Pragma']        = 'no-cache'
            resp.headers['Expires']       = '0'
        return resp

    # ── Auth decorators ────────────────────────────────────────────
    def login_required(f):
        @wraps(f)
        def inner(*a, **kw):
            if 'uid' not in session:
                return redirect(url_for('login'))
            return f(*a, **kw)
        return inner

    def admin_required(f):
        @wraps(f)
        def inner(*a, **kw):
            if 'uid' not in session:
                return redirect(url_for('login'))
            if session.get('role') != 'admin':
                flash('Acceso denegado.', 'danger')
                return redirect(url_for('dashboard'))
            return f(*a, **kw)
        return inner

    # ── Helpers de búsqueda ────────────────────────────────────────
    def _get_search(sid, require_owner=True):
        """Devuelve una búsqueda; admin ve todas, usuario solo las suyas."""
        if session.get('role') == 'admin':
            return db().execute(
                "SELECT s.*, u.name as u_name, u.email as u_email "
                "FROM searches s JOIN users u ON s.user_id=u.id WHERE s.id=?", (sid,)
            ).fetchone()
        if require_owner:
            return db().execute(
                "SELECT * FROM searches WHERE id=? AND user_id=?",
                (sid, session['uid'])
            ).fetchone()
        return db().execute("SELECT * FROM searches WHERE id=?", (sid,)).fetchone()

    # ══════════════════════════════════════════════════════════════
    # RUTAS AUTH
    # ══════════════════════════════════════════════════════════════
    @app.route('/')
    def index():
        return redirect(url_for('dashboard') if 'uid' in session else url_for('login'))

    @app.route('/login', methods=['GET', 'POST'])
    def login():
        if request.method == 'POST':
            email = request.form.get('email', '').strip().lower()
            pw    = request.form.get('password', '')
            u     = db().execute(
                "SELECT * FROM users WHERE email=? AND active=1", (email,)
            ).fetchone()
            if u and check_password_hash(u['password_hash'], pw):
                session.clear()
                session.update(uid=u['id'], uname=u['name'], role=u['role'])
                return redirect(url_for('dashboard'))
            flash('Correo o contraseña incorrectos.', 'danger')
        return render_template('login.html')

    @app.route('/logout')
    def logout():
        session.clear()
        return redirect(url_for('login'))

    @app.route('/profile', methods=['GET', 'POST'])
    @login_required
    def profile():
        d   = db()
        uid = session['uid']
        u   = d.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()
        cfg = {r['key']: r['value'] for r in d.execute("SELECT key,value FROM settings")}

        if request.method == 'POST':
            name       = request.form.get('name', '').strip()
            email      = request.form.get('email', '').strip().lower()
            pw         = request.form.get('password', '')
            pw2        = request.form.get('confirm', '')
            tg_chat_id = request.form.get('tg_chat_id', '').strip()

            if not name or not email:
                flash('Nombre y correo son obligatorios.', 'danger')
            elif d.execute("SELECT 1 FROM users WHERE email=? AND id!=?", (email, uid)).fetchone():
                flash('Ese correo ya está en uso.', 'danger')
            elif pw and pw != pw2:
                flash('Las contraseñas no coinciden.', 'danger')
            elif pw and len(pw) < 6:
                flash('Mínimo 6 caracteres en la contraseña.', 'danger')
            else:
                if pw:
                    d.execute("UPDATE users SET name=?,email=?,password_hash=?,tg_chat_id=? WHERE id=?",
                              (name, email, generate_password_hash(pw), tg_chat_id, uid))
                else:
                    d.execute("UPDATE users SET name=?,email=?,tg_chat_id=? WHERE id=?",
                              (name, email, tg_chat_id, uid))
                d.commit()
                session['uname'] = name
                flash('Perfil actualizado.', 'success')
                return redirect(url_for('profile'))

        return render_template('profile.html', u=u,
                               tg_token=cfg.get('tg_token', ''),
                               bot_name=cfg.get('tg_bot_name', ''))

    @app.route('/profile/test_telegram', methods=['POST'])
    @login_required
    def profile_test_telegram():
        from alerts.telegram import test_connection
        d       = db()
        uid     = session['uid']
        u       = d.execute("SELECT tg_chat_id FROM users WHERE id=?", (uid,)).fetchone()
        cfg     = {r['key']: r['value'] for r in d.execute("SELECT key,value FROM settings")}
        token   = cfg.get('tg_token', '')
        chat_id = u['tg_chat_id'] if u else ''
        if not token or not chat_id:
            return jsonify(ok=False, error='Bot no configurado o Chat ID vacío.')
        ok, err = test_connection(token, chat_id)
        return jsonify(ok=ok, error=err)

    @app.route('/register', methods=['GET', 'POST'])
    def register():
        if request.method == 'POST':
            name = request.form.get('name', '').strip()
            email= request.form.get('email', '').strip().lower()
            pw   = request.form.get('password', '')
            pw2  = request.form.get('confirm', '')
            if not all([name, email, pw]):
                flash('Todos los campos son obligatorios.', 'danger')
            elif pw != pw2:
                flash('Las contraseñas no coinciden.', 'danger')
            elif len(pw) < 6:
                flash('Mínimo 6 caracteres en la contraseña.', 'danger')
            elif db().execute("SELECT 1 FROM users WHERE email=?", (email,)).fetchone():
                flash('Ese correo ya está registrado.', 'danger')
            else:
                n    = db().execute("SELECT COUNT(*) FROM users").fetchone()[0]
                role = 'admin' if n == 0 else 'user'
                db().execute(
                    "INSERT INTO users (name,email,password_hash,role) VALUES (?,?,?,?)",
                    (name, email, generate_password_hash(pw), role)
                )
                db().commit()
                flash('Cuenta creada. Inicia sesión.', 'success')
                return redirect(url_for('login'))
        return render_template('register.html')

    # ══════════════════════════════════════════════════════════════
    # DASHBOARD
    # ══════════════════════════════════════════════════════════════
    @app.route('/dashboard')
    @login_required
    def dashboard():
        d   = db()
        uid = session['uid']
        is_admin = session['role'] == 'admin'

        if is_admin:
            searches = d.execute("""
                SELECT s.*, u.name as u_name,
                       (SELECT COUNT(*) FROM matches m WHERE m.search_id=s.id) as mc
                FROM searches s JOIN users u ON s.user_id=u.id
                ORDER BY CASE s.status
                           WHEN 'active'    THEN 0
                           WHEN 'paused'    THEN 1
                           WHEN 'completed' THEN 2
                           ELSE 3 END,
                         s.created_at DESC
            """).fetchall()
            total_users = d.execute("SELECT COUNT(*) FROM users").fetchone()[0]
            recent = d.execute("""
                SELECT m.*, s.name as s_name, s.id as s_id, u.name as u_name
                FROM matches m
                JOIN searches s ON m.search_id=s.id
                JOIN users    u ON s.user_id=u.id
                ORDER BY m.found_at DESC LIMIT 25
            """).fetchall()
        else:
            searches = d.execute("""
                SELECT s.*,
                       (SELECT COUNT(*) FROM matches m WHERE m.search_id=s.id) as mc
                FROM searches s WHERE s.user_id=?
                ORDER BY CASE s.status
                           WHEN 'active'    THEN 0
                           WHEN 'paused'    THEN 1
                           WHEN 'completed' THEN 2
                           ELSE 3 END,
                         s.created_at DESC
            """, (uid,)).fetchall()
            total_users = None
            recent = d.execute("""
                SELECT m.*, s.name as s_name, s.id as s_id
                FROM matches m JOIN searches s ON m.search_id=s.id
                WHERE s.user_id=? ORDER BY m.found_at DESC LIMIT 25
            """, (uid,)).fetchall()

        today_iso = date.today().isoformat()
        return render_template('dashboard.html',
            searches=searches,
            total_searches=len(searches),
            active_searches=sum(1 for s in searches
                                if s['status'] == 'active' and s['date_end'] >= today_iso),
            total_matches=sum(s['mc'] for s in searches),
            total_users=total_users,
            recent=recent,
        )

    # ══════════════════════════════════════════════════════════════
    # CONSULTAS IA (RAG)
    # ══════════════════════════════════════════════════════════════
    @app.route('/ask')
    @login_required
    def ask_page():
        from rag import RANGOS
        # Canales distintos visibles para el selector
        canales = []
        try:
            conn = sqlite3.connect(str(TRANS_DB), timeout=2)
            canales = [r[0] for r in conn.execute(
                "SELECT DISTINCT channel_name FROM transcriptions "
                "WHERE channel_name IS NOT NULL ORDER BY channel_name").fetchall()]
            conn.close()
        except Exception:
            pass
        return render_template('ask.html', rangos=RANGOS, canales=canales)

    @app.route('/api/ask', methods=['POST'])
    @login_required
    def api_ask():
        import json as _json
        from flask import Response, stream_with_context
        from rag   import ask_stream

        try:
            data = request.get_json(force=True) or {}
        except Exception:
            data = {}
        question = (data.get('q') or '').strip()
        rango    = data.get('rango', '24h')
        canal    = (data.get('canal') or '').strip() or None
        try:
            top_n = int(data.get('top_n') or 15)
        except Exception:
            top_n = 15

        if not question:
            return jsonify({'error': 'pregunta vacía'}), 400

        @stream_with_context
        def gen():
            for evt in ask_stream(question, rango=rango, canal=canal, top_n=top_n):
                yield _json.dumps(evt) + "\n"

        return Response(gen(), mimetype='application/x-ndjson')

    # ══════════════════════════════════════════════════════════════
    # BÚSQUEDAS — CRUD
    # ══════════════════════════════════════════════════════════════
    @app.route('/searches/new', methods=['GET', 'POST'])
    @login_required
    def search_new():
        if request.method == 'POST':
            name          = request.form.get('name', '').strip()
            kw_raw        = request.form.get('keywords', '').strip()
            phonetic      = 1 if request.form.get('phonetic') else 0
            whole_word    = 1 if request.form.get('whole_word') else 0
            d_start       = request.form.get('date_start', '')
            d_end         = request.form.get('date_end', '')
            dmode         = request.form.get('delivery_mode', 'final')
            remail        = request.form.get('report_email', '').strip()
            notify_tg     = 1 if request.form.get('notify_telegram') else 0

            if not all([name, kw_raw, d_start, d_end]):
                flash('Nombre, palabras y fechas son obligatorios.', 'danger')
            else:
                kws = [k.strip() for k in re.split(r'[\n,]+', kw_raw) if k.strip()]
                db().execute("""
                    INSERT INTO searches
                      (user_id,name,keywords,phonetic,whole_word,date_start,date_end,
                       delivery_mode,report_email,status,notify_telegram)
                    VALUES (?,?,?,?,?,?,?,?,?,'active',?)
                """, (session['uid'], name, json.dumps(kws, ensure_ascii=False),
                      phonetic, whole_word, d_start, d_end, dmode, remail, notify_tg))
                db().commit()
                flash(f'Búsqueda «{name}» creada.', 'success')
                return redirect(url_for('dashboard'))
        return render_template('search_new.html', today=date.today().isoformat())

    def _match_where(sid, kws, chs, pfs, date_from, date_to):
        """Construye WHERE + params para la tabla matches con todos los filtros activos."""
        conds, params = ['search_id=?'], [sid]
        if kws:
            conds.append(f"keyword IN ({','.join('?'*len(kws))})")
            params.extend(kws)
        if chs:
            conds.append(f"channel_name IN ({','.join('?'*len(chs))})")
            params.extend(chs)
        if date_from:
            conds.append("date(timestamp) >= ?"); params.append(date_from)
        if date_to:
            conds.append("date(timestamp) <= ?"); params.append(date_to)
        if pfs:
            conds.append(f"""EXISTS (
                SELECT 1 FROM epg_programmes e
                WHERE e.channel_name = matches.channel_name
                  AND e.start_ts <= matches.timestamp
                  AND e.stop_ts  >  matches.timestamp
                  AND e.title IN ({','.join('?'*len(pfs))})
            )""")
            params.extend(pfs)
        return ' AND '.join(conds), params

    @app.route('/searches/<int:sid>')
    @login_required
    def search_detail(sid):
        s = _get_search(sid)
        if not s:
            flash('Búsqueda no encontrada.', 'danger')
            return redirect(url_for('dashboard'))

        page      = request.args.get('page', 1, type=int)
        pp        = 50
        kfs       = request.args.getlist('kw')
        cfs       = request.args.getlist('ch')
        pfs       = request.args.getlist('prog')
        date_from = request.args.get('date_from', '')
        date_to   = request.args.get('date_to', '')

        where, params = _match_where(sid, kfs, cfs, pfs, date_from, date_to)

        d         = db()
        total     = d.execute(f"SELECT COUNT(*) FROM matches WHERE {where}", params).fetchone()[0]
        total_all = d.execute("SELECT COUNT(*) FROM matches WHERE search_id=?", (sid,)).fetchone()[0]
        matches_raw = d.execute(
            f"SELECT * FROM matches WHERE {where} ORDER BY found_at DESC LIMIT ? OFFSET ?",
            params + [pp, (page-1)*pp]
        ).fetchall()
        # Enriquece cada match con precise_timestamp y centered_text
        matches = [_enrich_match(m, phonetic=bool(s['phonetic']), whole_word=bool(s['whole_word']))
                   for m in matches_raw]
        kw_all = d.execute("SELECT DISTINCT keyword FROM matches WHERE search_id=? ORDER BY keyword", (sid,)).fetchall()
        ch_all = d.execute("SELECT DISTINCT channel_name FROM matches WHERE search_id=? ORDER BY channel_name", (sid,)).fetchall()
        prog_all = d.execute("""
            SELECT DISTINCT e.title
            FROM matches m
            JOIN epg_programmes e
              ON e.channel_name = m.channel_name
             AND e.start_ts <= m.timestamp
             AND e.stop_ts  >  m.timestamp
            WHERE m.search_id = ?
            ORDER BY e.title
            LIMIT 200
        """, (sid,)).fetchall()
        kw_stats = d.execute(
            "SELECT keyword, COUNT(*) cnt FROM matches WHERE search_id=? GROUP BY keyword ORDER BY cnt DESC", (sid,)
        ).fetchall()
        ch_stats = d.execute(
            "SELECT channel_name, COUNT(*) cnt FROM matches WHERE search_id=? GROUP BY channel_name ORDER BY cnt DESC", (sid,)
        ).fetchall()

        # Heatmap: una fila por fecha real (date_start … hoy o date_end)
        hm_start = date.fromisoformat(s['date_start'])
        hm_end   = min(date.today(), date.fromisoformat(s['date_end']))
        if hm_end < hm_start:
            hm_end = hm_start
        hm_dates_list = []
        cur = hm_start
        while cur <= hm_end:
            hm_dates_list.append(cur.isoformat())
            cur += timedelta(days=1)
        n_days = len(hm_dates_list)
        date_idx = {dt: i for i, dt in enumerate(hm_dates_list)}

        hm_rows = d.execute("""
            SELECT date(timestamp) as day,
                   CAST(strftime('%H', timestamp) AS INTEGER) as hr,
                   COUNT(*) as cnt
            FROM matches WHERE search_id=?
            GROUP BY day, hr
        """, (sid,)).fetchall()
        heatmap = [[0] * 24 for _ in range(n_days)]
        for r in hm_rows:
            idx = date_idx.get(r['day'])
            if idx is not None:
                heatmap[idx][r['hr']] = r['cnt']
        hm_max = max((heatmap[di][h] for di in range(n_days) for h in range(24)), default=0)

        # EPG: obtener programa para cada coincidencia en esta página
        from alerts.epg import get_programme_at
        prog_map = {}
        for m in matches:
            if m['timestamp']:
                prog_map[m['id']] = get_programme_at(d, m['channel_name'] or '', m['timestamp'])

        return render_template('search_detail.html',
            s=s, keywords=json.loads(s['keywords']),
            matches=matches, total=total, page=page, pp=pp,
            pages=max(1, (total-1)//pp+1),
            kw_all=kw_all, ch_all=ch_all, prog_all=prog_all,
            kfs=kfs, cfs=cfs, pfs=pfs,
            date_from=date_from, date_to=date_to,
            total_all=total_all,
            kw_stats=kw_stats, ch_stats=ch_stats,
            heatmap=heatmap, hm_max=hm_max, hm_dates=hm_dates_list,
            prog_map=prog_map,
        )

    def _match_moment(sid, mid):
        """Devuelve (match_row, datetime del instante exacto de la palabra) o (None, None)."""
        s = _get_search(sid)
        if not s:
            return None, None
        m = db().execute(
            "SELECT * FROM matches WHERE id=? AND search_id=?", (mid, sid)
        ).fetchone()
        if not m:
            return None, None
        md = _enrich_match(m, phonetic=bool(s['phonetic']), whole_word=bool(s['whole_word']))
        try:
            moment = datetime.fromisoformat(md['precise_timestamp'])
        except Exception:
            return m, None
        return m, moment

    @app.route('/searches/<int:sid>/matches/<int:mid>/snapshot.jpg')
    @login_required
    def match_snapshot(sid, mid):
        from alerts.clips import extract_snapshot, CACHE_DIR
        m, moment = _match_moment(sid, mid)
        if not m or not moment:
            return ('', 404)
        out = CACHE_DIR / f'{mid}.jpg'
        if not out.exists():
            if not extract_snapshot(m['channel_name'] or '', moment, out):
                return ('', 404)
        return send_file(out, mimetype='image/jpeg')

    @app.route('/searches/<int:sid>/matches/<int:mid>/clip.mp4')
    @login_required
    def match_clip(sid, mid):
        from alerts.clips import extract_clip, CACHE_DIR
        m, moment = _match_moment(sid, mid)
        if not m or not moment:
            return ('', 404)
        out = CACHE_DIR / f'{mid}.mp4'
        if not out.exists():
            if not extract_clip(m['channel_name'] or '', moment, out):
                return ('', 404)
        return send_file(out, mimetype='video/mp4')

    # ══════════════════════════════════════════════════════════════
    # MONITOR DE SEÑALES (mosaico en vivo)
    # ══════════════════════════════════════════════════════════════
    @app.route('/videowall')
    @login_required
    def videowall():
        from alerts.videowall import list_channels
        return render_template('videowall.html', channels=list_channels())

    @app.route('/videowall/thumb/<int:num>.jpg')
    @login_required
    def videowall_thumb(num):
        from alerts.videowall import list_channels, get_thumbnail
        ch = next((c for c in list_channels() if c['num'] == num), None)
        if ch is None:
            return ('', 404)
        out = get_thumbnail(num, ch['folder'])
        if out is None:
            return ('', 404)
        resp = send_file(out, mimetype='image/jpeg')
        resp.headers['Cache-Control'] = 'no-store'
        return resp

    @app.route('/searches/<int:sid>/edit', methods=['GET', 'POST'])
    @login_required
    def search_edit(sid):
        s = _get_search(sid)
        if not s:
            flash('Búsqueda no encontrada.', 'danger')
            return redirect(url_for('dashboard'))

        if request.method == 'POST':
            name      = request.form.get('name', '').strip()
            kw_raw    = request.form.get('keywords', '').strip()
            kws       = [k.strip() for k in re.split(r'[\n,]+', kw_raw) if k.strip()]
            notify_tg  = 1 if request.form.get('notify_telegram') else 0
            new_start  = request.form.get('date_start')
            new_end    = request.form.get('date_end')
            new_kws    = json.dumps(kws, ensure_ascii=False)
            new_phon   = 1 if request.form.get('phonetic') else 0
            new_whole  = 1 if request.form.get('whole_word') else 0

            # Si cambian fechas, palabras o tipo de búsqueda → re-escanear histórico
            needs_reinit = (
                new_start != s['date_start']  or
                new_end   != s['date_end']    or
                new_kws   != s['keywords']    or
                new_phon  != s['phonetic']    or
                new_whole != s['whole_word']
            )

            db().execute("""
                UPDATE searches SET
                  name=?, keywords=?, phonetic=?, whole_word=?, date_start=?, date_end=?,
                  delivery_mode=?, report_email=?, status=?, notify_telegram=?,
                  initialized=?
                WHERE id=?
            """, (name, new_kws, new_phon, new_whole, new_start, new_end,
                  request.form.get('delivery_mode', 'final'),
                  request.form.get('report_email', '').strip(),
                  request.form.get('status', 'active'), notify_tg,
                  0 if needs_reinit else s['initialized'],
                  sid))

            if needs_reinit:
                db().execute("DELETE FROM matches WHERE search_id=?", (sid,))
                flash('Búsqueda actualizada. Re-escaneando histórico con el nuevo rango…', 'success')
            else:
                flash('Búsqueda actualizada.', 'success')

            db().commit()
            return redirect(url_for('search_detail', sid=sid))

        return render_template('search_edit.html', s=s,
                               keywords=json.loads(s['keywords']))

    @app.route('/searches/<int:sid>/toggle', methods=['POST'])
    @login_required
    def search_toggle(sid):
        s = _get_search(sid)
        if s:
            ns = 'paused' if s['status'] == 'active' else 'active'
            db().execute("UPDATE searches SET status=? WHERE id=?", (ns, sid))
            db().commit()
        return redirect(request.referrer or url_for('dashboard'))

    @app.route('/searches/<int:sid>/delete', methods=['POST'])
    @login_required
    def search_delete(sid):
        s = _get_search(sid)
        if s:
            db().execute("DELETE FROM matches  WHERE search_id=?", (sid,))
            db().execute("DELETE FROM searches WHERE id=?",        (sid,))
            db().commit()
            flash('Búsqueda eliminada.', 'success')
        return redirect(url_for('dashboard'))

    @app.route('/searches/<int:sid>/report', methods=['POST'])
    @login_required
    def search_report(sid):
        from alerts.mailer import send_report
        s = _get_search(sid)
        if not s:
            flash('Búsqueda no encontrada.', 'danger')
            return redirect(url_for('dashboard'))
        kfs       = request.form.getlist('kw')
        cfs       = request.form.getlist('ch')
        pfs       = request.form.getlist('prog')
        date_from = request.form.get('date_from', '')
        date_to   = request.form.get('date_to', '')
        where, params = _match_where(sid, kfs, cfs, pfs, date_from, date_to)
        d       = db()
        matches = d.execute(f"SELECT * FROM matches WHERE {where} ORDER BY found_at DESC", params).fetchall()
        rows    = d.execute("SELECT key,value FROM settings").fetchall()
        cfg     = {r['key']: r['value'] for r in rows}
        if not cfg.get('smtp_host'):
            flash('Configura el servidor SMTP primero.', 'warning')
        else:
            ok, msg = send_report(s, [dict(m) for m in matches], cfg, 'manual')
            flash(msg, 'success' if ok else 'danger')
        return redirect(url_for('search_detail', sid=sid))

    @app.route('/searches/<int:sid>/report_telegram', methods=['POST'])
    @login_required
    def search_report_telegram(sid):
        from alerts.telegram import send_report as tg_send_report
        s = _get_search(sid)
        if not s:
            flash('Búsqueda no encontrada.', 'danger')
            return redirect(url_for('dashboard'))
        kfs       = request.form.getlist('kw')
        cfs       = request.form.getlist('ch')
        pfs       = request.form.getlist('prog')
        date_from = request.form.get('date_from', '')
        date_to   = request.form.get('date_to', '')
        where, params = _match_where(sid, kfs, cfs, pfs, date_from, date_to)
        d       = db()
        matches = d.execute(f"SELECT * FROM matches WHERE {where} ORDER BY found_at DESC", params).fetchall()
        cfg     = {r['key']: r['value'] for r in d.execute("SELECT key,value FROM settings")}
        token   = cfg.get('tg_token', '')
        # Usa el chat_id personal del dueño de la búsqueda; fallback al global
        owner   = d.execute("SELECT tg_chat_id FROM users WHERE id=?", (s['user_id'],)).fetchone()
        chat_id = (owner['tg_chat_id'] if owner else '') or cfg.get('tg_chat_id', '')
        if not token or not chat_id:
            flash('Configura tu Chat ID de Telegram en Mi perfil primero.', 'warning')
        else:
            ok, err = tg_send_report(token, chat_id, dict(s), [dict(m) for m in matches])
            flash('Reporte enviado a Telegram.' if ok else f'Error Telegram: {err}',
                  'success' if ok else 'danger')
        return redirect(url_for('search_detail', sid=sid))

    @app.route('/settings/test_telegram', methods=['POST'])
    @admin_required
    def settings_test_telegram():
        from alerts.telegram import test_connection
        d       = db()
        cfg     = {r['key']: r['value'] for r in d.execute("SELECT key,value FROM settings")}
        token   = cfg.get('tg_token', '')
        chat_id = cfg.get('tg_chat_id', '')
        if not token or not chat_id:
            flash('Ingresa el Token y Chat ID antes de probar.', 'warning')
        else:
            ok, err = test_connection(token, chat_id)
            flash('Mensaje de prueba enviado correctamente.' if ok else f'Error: {err}',
                  'success' if ok else 'danger')
        return redirect(url_for('settings'))

    # ══════════════════════════════════════════════════════════════
    # ADMIN
    # ══════════════════════════════════════════════════════════════
    @app.route('/admin')
    @admin_required
    def admin():
        d = db()
        users = d.execute("""
            SELECT u.*,
                   (SELECT COUNT(*) FROM searches s WHERE s.user_id=u.id) sc
            FROM users u ORDER BY u.created_at DESC
        """).fetchall()
        searches = d.execute("""
            SELECT s.*, u.name as u_name,
                   (SELECT COUNT(*) FROM matches m WHERE m.search_id=s.id) mc
            FROM searches s JOIN users u ON s.user_id=u.id
            ORDER BY s.created_at DESC
        """).fetchall()
        return render_template('admin.html', users=users, searches=searches)

    @app.route('/admin/users/<int:uid>/toggle', methods=['POST'])
    @admin_required
    def admin_user_toggle(uid):
        if uid != session['uid']:
            u = db().execute("SELECT active FROM users WHERE id=?", (uid,)).fetchone()
            if u:
                db().execute("UPDATE users SET active=? WHERE id=?",
                             (0 if u['active'] else 1, uid))
                db().commit()
        return redirect(url_for('admin'))

    @app.route('/admin/users/<int:uid>/role', methods=['POST'])
    @admin_required
    def admin_user_role(uid):
        if uid != session['uid']:
            role = request.form.get('role', 'user')
            db().execute("UPDATE users SET role=? WHERE id=?", (role, uid))
            db().commit()
        return redirect(url_for('admin'))

    # ══════════════════════════════════════════════════════════════
    # SETTINGS (SMTP)
    # ══════════════════════════════════════════════════════════════
    @app.route('/settings', methods=['GET', 'POST'])
    @admin_required
    def settings():
        from alerts.epg import get_coverage_stats
        d = db()
        if request.method == 'POST':
            for k in ['smtp_host','smtp_port','smtp_user','smtp_pass','smtp_from','smtp_tls',
                      'tg_token','tg_chat_id']:
                d.execute("INSERT OR REPLACE INTO settings (key,value) VALUES (?,?)",
                          (k, request.form.get(k, '')))
            d.commit()
            flash('Configuración guardada.', 'success')
        cfg       = {r['key']: r['value'] for r in d.execute("SELECT key,value FROM settings")}
        epg_stats = get_coverage_stats(d)
        return render_template('settings.html', cfg=cfg, epg_stats=epg_stats)

    @app.route('/settings/epg_fetch', methods=['POST'])
    @admin_required
    def settings_epg_fetch():
        from alerts.epg import fetch_current
        n = fetch_current()
        db().execute("INSERT OR REPLACE INTO settings (key,value) VALUES ('epg_last_fetch',?)",
                     (datetime.now().strftime('%Y-%m-%d %H:%M:%S'),))
        db().commit()
        flash(f'EPG actualizado: {n} programas nuevos guardados.', 'success')
        return redirect(url_for('settings'))

    # ══════════════════════════════════════════════════════════════
    # API JSON (para auto-refresh)
    # ══════════════════════════════════════════════════════════════
    @app.route('/api/searches/<int:sid>/matches')
    @login_required
    def api_matches(sid):
        s = _get_search(sid)
        if not s:
            return jsonify(error='not found'), 404
        d        = db()
        since    = request.args.get('since', '')
        after_id = request.args.get('after_id', type=int)
        if after_id is not None:
            rows = d.execute(
                "SELECT * FROM matches WHERE search_id=? AND id>? ORDER BY id ASC LIMIT 100",
                (sid, after_id)
            ).fetchall()
        elif since:
            rows = d.execute(
                "SELECT * FROM matches WHERE search_id=? AND found_at>? ORDER BY found_at DESC LIMIT 50",
                (sid, since)
            ).fetchall()
        else:
            rows = d.execute(
                "SELECT * FROM matches WHERE search_id=? ORDER BY found_at DESC LIMIT 50", (sid,)
            ).fetchall()
        total = d.execute("SELECT COUNT(*) FROM matches WHERE search_id=?", (sid,)).fetchone()[0]
        phonetic   = bool(s['phonetic'])
        whole_word = bool(s['whole_word'])
        return jsonify(matches=[_enrich_match(r, phonetic=phonetic, whole_word=whole_word) for r in rows],
                       total=total)

    @app.route('/api/stats')
    @login_required
    def api_stats():
        d        = db()
        uid      = session['uid']
        is_admin = session.get('role') == 'admin'

        if is_admin:
            searches = d.execute("""
                SELECT s.id, s.name, s.status,
                       (SELECT COUNT(*) FROM matches m WHERE m.search_id=s.id) as mc
                FROM searches s ORDER BY s.created_at DESC
            """).fetchall()
            recent = d.execute("""
                SELECT m.id, m.keyword, m.channel_name, m.timestamp, m.matched_text,
                       s.name as s_name, s.id as s_id
                FROM matches m
                JOIN searches s ON m.search_id=s.id
                ORDER BY m.found_at DESC LIMIT 25
            """).fetchall()
            total_matches = d.execute("SELECT COUNT(*) FROM matches").fetchone()[0]
        else:
            searches = d.execute("""
                SELECT s.id, s.name, s.status,
                       (SELECT COUNT(*) FROM matches m WHERE m.search_id=s.id) as mc
                FROM searches s WHERE s.user_id=? ORDER BY s.created_at DESC
            """, (uid,)).fetchall()
            recent = d.execute("""
                SELECT m.id, m.keyword, m.channel_name, m.timestamp, m.matched_text,
                       s.name as s_name, s.id as s_id
                FROM matches m JOIN searches s ON m.search_id=s.id
                WHERE s.user_id=? ORDER BY m.found_at DESC LIMIT 25
            """, (uid,)).fetchall()
            total_matches = sum(s['mc'] for s in searches)

        active_searches = sum(1 for s in searches if s['status'] == 'active')

        return jsonify(
            searches=[{'id': s['id'], 'mc': s['mc']} for s in searches],
            total_matches=total_matches,
            active_searches=active_searches,
            recent=[{
                'id':           m['id'],
                'keyword':      m['keyword'],
                'channel_name': m['channel_name'],
                'timestamp':    m['timestamp'],
                'matched_text': m['matched_text'],
                's_name':       m['s_name'],
                's_id':         m['s_id'],
            } for m in recent],
        )

    @app.route('/api/searches/<int:sid>/progress')
    @login_required
    def api_search_progress(sid):
        d   = db()
        uid = session['uid']
        is_admin = session.get('role') == 'admin'
        row = d.execute(
            "SELECT initialized, init_rows_done, init_rows_total FROM searches WHERE id=?"
            + (" AND (user_id=? OR 1=?)" if not is_admin else ""),
            (sid, uid, 1) if not is_admin else (sid,)
        ).fetchone()
        if not row:
            return jsonify(error='not found'), 404
        done  = row['init_rows_done']  or 0
        total = row['init_rows_total'] or 0
        pct   = round(done / total * 100) if total > 0 else (100 if row['initialized'] else 0)
        return jsonify(
            initialized = bool(row['initialized']),
            done        = done,
            total       = total,
            pct         = pct,
        )

    # ══════════════════════════════════════════════════════════════
    # EXPORTAR A EXCEL
    # ══════════════════════════════════════════════════════════════
    @app.route('/export', methods=['POST'])
    @login_required
    def export():
        import io
        import openpyxl
        from openpyxl.styles          import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils           import get_column_letter
        from openpyxl.cell.rich_text  import CellRichText, TextBlock
        from openpyxl.cell.text       import InlineFont

        search_ids = request.form.getlist('search_ids')
        search_ids = [int(x) for x in search_ids if x.isdigit()]
        if not search_ids:
            flash('Selecciona al menos una búsqueda.', 'warning')
            return redirect(url_for('dashboard'))
        # Filtros opcionales (solo aplican si viene de una búsqueda individual)
        _exp_kfs       = request.form.getlist('kw')
        _exp_cfs       = request.form.getlist('ch')
        _exp_pfs       = request.form.getlist('prog')
        _exp_date_from = request.form.get('date_from', '')
        _exp_date_to   = request.form.get('date_to', '')
        _exp_single    = len(search_ids) == 1   # filtros solo para exportación individual

        d   = db()
        tdb = sqlite3.connect(str(TRANS_DB), timeout=10)
        tdb.row_factory = sqlite3.Row
        tdb.execute("PRAGMA journal_mode=WAL")

        from alerts.epg import get_programme_at

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # Estilos
        h_font   = Font(bold=True, color='FFFFFF', size=11, name='Calibri')
        h_fill   = PatternFill('solid', fgColor='1D4ED8')
        h_align  = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin     = Side(style='thin', color='CBD5E1')
        border   = Border(left=thin, right=thin, top=thin, bottom=thin)
        alt_fill = PatternFill('solid', fgColor='F1F5F9')
        hl_fill  = PatternFill('solid', fgColor='FEF9C3')   # amarillo suave para fila activa

        def _bold_kw(text, keyword, phonetic=False, whole_word=False):
            """Devuelve CellRichText con la keyword en negrita.
            Modo exacto: case-insensitive. Modo fonético: resalta cada palabra
            del texto que sea fonéticamente equivalente al keyword. En modo
            whole_word exige límites de palabra (no resalta "día" dentro de
            "diálogo")."""
            if not text or not keyword:
                return text or ''
            bold = InlineFont(b=True, color='1D4ED8')
            parts, last = [], 0
            if phonetic:
                ph_kw   = _phonetic(keyword)
                pattern = re.compile(r'(?<!\w)' + re.escape(ph_kw) + r'(?!\w)') if whole_word else None
                for m in re.finditer(r'\S+', text):
                    word    = m.group()
                    word_ph = _phonetic(word)
                    matched = bool(pattern.search(word_ph)) if whole_word else (ph_kw in word_ph)
                    if matched:
                        if m.start() > last:
                            parts.append(text[last:m.start()])
                        parts.append(TextBlock(bold, word))
                        last = m.end()
            else:
                kw_pattern = re.escape(keyword)
                if whole_word:
                    kw_pattern = r'(?<!\w)' + kw_pattern + r'(?!\w)'
                for m in re.compile(kw_pattern, re.IGNORECASE).finditer(text):
                    if m.start() > last:
                        parts.append(text[last:m.start()])
                    parts.append(TextBlock(bold, m.group()))
                    last = m.end()
            if last < len(text):
                parts.append(text[last:])
            return CellRichText(*parts) if parts else text

        WINDOW_SEC = 15

        for sid in search_ids:
            s = _get_search(sid)
            if not s:
                continue

            if _exp_single and any([_exp_kfs, _exp_cfs, _exp_pfs, _exp_date_from, _exp_date_to]):
                _w, _p = _match_where(sid, _exp_kfs, _exp_cfs, _exp_pfs, _exp_date_from, _exp_date_to)
                matches = d.execute(
                    f"SELECT * FROM matches WHERE {_w} ORDER BY timestamp ASC", _p
                ).fetchall()
            else:
                matches = d.execute(
                    "SELECT * FROM matches WHERE search_id=? ORDER BY timestamp ASC", (sid,)
                ).fetchall()

            # ── Precarga de contexto: 1 query por canal en vez de 1 por fila ──
            channel_ranges = {}
            for m in matches:
                cid, ts = m['channel_id'], m['timestamp']
                if cid and ts:
                    lo, hi = channel_ranges.get(cid, (ts, ts))
                    channel_ranges[cid] = (min(lo, ts), max(hi, ts))

            ctx_data = {}   # channel_id -> [(timestamp_str, text), ...]
            for cid, (ts_min, ts_max) in channel_ranges.items():
                trans_rows = tdb.execute("""
                    SELECT timestamp, text FROM transcriptions
                    WHERE channel_id = ?
                      AND timestamp >= datetime(?, ?)
                      AND timestamp <= datetime(?, ?)
                      AND text IS NOT NULL AND text != '[~]'
                    ORDER BY timestamp ASC
                """, (cid,
                      ts_min, f'-{WINDOW_SEC} seconds',
                      ts_max, f'+{WINDOW_SEC} seconds')).fetchall()
                ctx_data[cid] = [(r['timestamp'], r['text']) for r in trans_rows]

            def _get_context(channel_id, timestamp):
                if not channel_id or not timestamp:
                    return ''
                entries = ctx_data.get(channel_id, [])
                if not entries:
                    return ''
                dt = datetime.strptime(timestamp[:19], '%Y-%m-%d %H:%M:%S')
                lo = (dt - timedelta(seconds=WINDOW_SEC)).strftime('%Y-%m-%d %H:%M:%S')
                hi = (dt + timedelta(seconds=WINDOW_SEC)).strftime('%Y-%m-%d %H:%M:%S')
                return ' '.join(t for ts2, t in entries if lo <= ts2 <= hi)

            sheet_name = re.sub(r'[\\/*?:\[\]]', '', s['name'])[:31] or f'Busqueda_{sid}'
            ws = wb.create_sheet(title=sheet_name)

            # ── Encabezado ──
            ws.merge_cells('A1:G1')
            c = ws['A1']
            c.value     = f"Monitoreo ITESO — {s['name']}"
            c.font      = Font(bold=True, size=13, color='1D4ED8', name='Calibri')
            c.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 22

            ws['A2'] = f"Período: {s['date_start']} → {s['date_end']}"
            ws['C2'] = f"Total coincidencias: {len(matches)}"
            ws['G2'] = f"Exportado: {date.today().isoformat()}"
            for cell in [ws['A2'], ws['C2'], ws['G2']]:
                cell.font = Font(italic=True, size=9, color='64748B', name='Calibri')
            ws.row_dimensions[2].height = 16
            ws.append([])

            # ── Cabeceras ──
            headers = ['Fecha / Hora Señal', 'Canal', 'Programa (EPG)',
                       'Palabra Detectada', 'Segmento detectado',
                       'Contexto ampliado (±15 seg)', 'Fecha Detección']
            ws.append(headers)
            hrow = ws.max_row
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=hrow, column=col)
                cell.font      = h_font
                cell.fill      = h_fill
                cell.alignment = h_align
                cell.border    = border
            ws.row_dimensions[hrow].height = 20

            # ── Datos ──
            for i, m in enumerate(matches):
                row_num  = hrow + 1 + i
                contexto = _get_context(m['channel_id'], m['timestamp'])
                programa = get_programme_at(d, m['channel_name'] or '', m['timestamp'] or '')

                ws.cell(row=row_num, column=1, value=str(m['timestamp'] or '')[:19])
                ws.cell(row=row_num, column=2, value=m['channel_name'] or '')
                ws.cell(row=row_num, column=3, value=programa)
                ws.cell(row=row_num, column=4, value=m['keyword'] or '')
                ws.cell(row=row_num, column=5, value=_bold_kw(m['matched_text'] or '', m['keyword'] or '', bool(s['phonetic']), bool(s['whole_word'])))
                ws.cell(row=row_num, column=6, value=_bold_kw(contexto, m['keyword'] or '', bool(s['phonetic']), bool(s['whole_word'])))
                ws.cell(row=row_num, column=7, value=str(m['found_at'] or '')[:19])

                fill = alt_fill if i % 2 == 0 else None
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_num, column=col)
                    cell.font      = Font(size=10, name='Calibri')
                    cell.border    = border
                    cell.alignment = Alignment(vertical='top',
                                               wrap_text=(col in (5, 6)))
                    if fill:
                        cell.fill = fill

                # Altura dinámica según longitud del contexto
                ctx_len = len(contexto)
                ws.row_dimensions[row_num].height = (
                    80 if ctx_len > 500 else
                    50 if ctx_len > 200 else
                    30 if ctx_len > 80  else 18
                )

            # ── Anchos ──
            ws.column_dimensions['A'].width = 22
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 50
            ws.column_dimensions['F'].width = 80
            ws.column_dimensions['G'].width = 22

            ws.freeze_panes = ws.cell(row=hrow + 1, column=1)

        tdb.close()

        if not wb.sheetnames:
            flash('No se encontraron búsquedas válidas.', 'warning')
            return redirect(url_for('dashboard'))

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"monitoreo_iteso_{date.today().isoformat()}.xlsx"
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename,
        )

    return app
